""" Zoopla scraping project """
# ---------------------------------------------- #
# Project starts at: 2020 02 01                  #
# Lithuana, by Vytautas Bielinskas               #
# ---------------------------------------------- #

# Import libraries
import requests, re, os
import pandas as pd
from bs4 import BeautifulSoup
import quick_functions as QF

""" Generate the list of URLs : Start"""
def generateURLs(pages):
    listURLs = []
    base_url = "https://www.zoopla.co.uk/to-rent/property/london/west-drayton/?identifier=london%2Fwest-drayton&q=West Drayton%2C London&search_source=refine&radius=0&price_frequency=per_month&pn="
    for i in range(1, pages+1, 1):
        fullURL = base_url + str(i)
        listURLs.append(fullURL)
    return listURLs
""" Generate the list of URLs : End"""

""" Get Data : Start"""
properties = [] # Here all the data will be stored
def getData(listURLs):
    
    import re
    from selenium import webdriver
    options = webdriver.chrome.options.Options()
    options.add_argument("--disable-extensions")
    
    featuresOfProperties = {}

    x = input('Debuging checkpoint.')
        
    def getFullDescription(objectURL):
            
        r = requests.get(objectURL)
        c = r.content
        soup = BeautifulSoup(c, "html.parser")
            
        try:
            description = soup.find("div", {"itemprop":"description"}).text
        except:
            description = ""
        cleaned = description.replace(",", " ")
        cleaned = cleaned.replace(":", " ")
        cleaned = cleaned.replace(".", " ")
        cleaned = cleaned.replace("W C", "WC")
        cleaned = cleaned.replace("/", " ")
        cleaned = cleaned.replace(")", ") ")
        cleaned = cleaned.replace("Description", "Description ")
        cleaned = cleaned.replace("Floor", "Floor ")
        cleaned = cleaned.replace("Floor", "Floor ")
        cleaned = cleaned.replace("Annexe", "")
        cleaned = cleaned.replace("Divided", " Divided")
        cleaned = cleaned.replace("  ", " ")
        return cleaned
    """ ------------------------------------------------------------------ """
    
    """ Parse stations : start """
    def singleStation(station):
        station = station.text
        station = station.replace(' ', '')
        station = station.replace('\n', '')
        station = station.split('(')[1].split(')')[0]
        station = station.replace('miles', '')
        print('--> singleStation: {}'.format(station))
        return float(station)
    """ Parse stations : end """
        
    for page in range(1, len(listURLs), 1): #(1, len(listURLs), 1):
        print(page," : ", listURLs[page])
        r = requests.get(listURLs[page])
        c = r.content
        soup = BeautifulSoup(c, "html.parser")
            
        infoTable = soup.find("ul", {"class":"listing-results"})
        items = infoTable.find_all("div", {"class":"listing-results-wrapper"})
            
        for item in range (0, len(items), 1):

            blockProperty = items[item]
            x = input('Debuging checkpoint.')
                    
            """ get Address """
            address = blockProperty.find("a", {"class":"listing-results-address"}).text
                    
            """ Listed on """
            try:
                listedOn = blockProperty.find("p", {"class":"top-half"}).find("small").text.replace('\n', '').split('Listed on  ')[1]
            except:
                listedOn = blockProperty.find("p", {"class":"top-half"}).find("small").text.replace('\n', '').split('Listed on ')[1]
            year_re = re.compile(r'\d\d\d\d')
            year = year_re.findall(listedOn)[0]
            listedOn = blockProperty.find("p", {"class":"top-half"}).find("small").text.split("\n")[2][1:].split(year)[0]
            listedOn = '{}{}'.format(listedOn, year)
                
            print('Listed on: {}'.format(listedOn))
            monthNum = " "
            if "Nov" in listedOn:
                monthNum = 11
            elif "Oct" in listedOn:
                monthNum = 10
            elif "Sep" in listedOn:
                monthNum = 9
            elif "Jun" in listedOn:
                monthNum = 6
            elif "Apr" in listedOn:
                monthNum = 4
            elif "Jan" in listedOn:
                monthNum = 1
            elif "Feb" in listedOn:
                monthNum = 2
            elif "Mar" in listedOn:
                monthNum = 3
            elif "May" in listedOn:
                monthNum = 5
            elif "Aug" in listedOn:
                monthNum = 8
            elif "Dec" in listedOn:
                monthNum = 12
            elif "Jul" in listedOn:
                monthNum = 7
            
            if "th" in listedOn:
                day = listedOn.split("th")[0]
            elif "nd" in listedOn:
                day = listedOn.split("nd")[0]
            elif "st" in listedOn:
                day = listedOn.split("st")[0]
            elif "rd" in listedOn:
                day = listedOn.split("rd")[0]
                
            year = "2020" 
            x = input('Debuging checkpoint.')
            
            listedOn = day + "/" + str(monthNum) + "/" + year
                    
            """ get ID """
            idPro = blockProperty.parent["data-listing-id"]
                    
            """ get Price"""
            pricePro_2 = []
            pricePro = blockProperty.find_all("", {"class":"listing-results-price"})[0].text.replace(" ", "")
            pricePro = pricePro.split("\n")
            for item in pricePro:
                if item:
                    pricePro_2.append(str(item))
            pricePro = pricePro_2[0].replace("£", "")
                    
            """ get Bedrooms """
            bedrooms = 0
            try:
                blockForAttrs = blockProperty.find("h3").find_all("span")
                for x in range(0, len(blockForAttrs), 1):
                    try:
                        if "num-beds" in blockForAttrs[x]["class"]:
                            bedrooms = float(blockForAttrs[x].text)
                    except:
                        print("No beds")
            except:
                bedrooms = 0
                        
            """ get Bathrooms """
            bathrooms = 0
            try:
                blockForAttrs = blockProperty.find("h3").find_all("span")
                bathrooms = 0
                for x in range(0, len(blockForAttrs), 1):
                    try:
                        if "num-baths" in blockForAttrs[x]["class"]:
                            bathrooms = float(blockForAttrs[x].text)
                    except:
                        print("No bath")
            except:
                bathrooms = 0
                        
            """ get Reception rooms """
            receptions = 0
            try:
                blockForAttrs = blockProperty.find("h3").find_all("span")
                receptions = 0
                for x in range(0, len(blockForAttrs), 1):
                    try:
                        if "num-reception" in blockForAttrs[x]["class"]:
                            receptions = float(blockForAttrs[x].text)
                    except:
                        print("No bath")
            except:
                receptions = 0
                    
            ppM = pricePro.split("pcm")[0].replace(",", "")
            
            """ get description """
            try:
                descriptionText = blockProperty.find_all("p")[-3].text.replace("\n", "").replace("  ", "")
                print("DESCRIPTION:", descriptionText)
            except:
                print("No description")
                descriptionText = "No description"
                
            if len(descriptionText) > 1:
                whiteIndex = 0
                z = 0
                while descriptionText[z] == " ":
                    whiteIndex = whiteIndex + 1
                    z = z + 1
                descriptionText = descriptionText[z:]
                
            """ get nearby station : working """
            nearbyStation = ""
            station_here = blockProperty.find("div", {"class":"nearby_stations_schools"}).text.split("\n")
            z = 0
            while len(station_here[z]) < 2:
                z = z + 1
            nearbyStation = station_here[z]
            
            """ IMPROVEMENT """
            nearby_stations = blockProperty.find('div', {'class':'nearby_stations_schools'}).find_all('li')
            print('The size of stations: {}.'.format(len(nearby_stations)))
            
            number_of_stations = len(nearby_stations)
            
            # Build float array for distances
            min_dist = 0   # Before evaluation
            try:
                stations_fl = []
                for this_station in range(0, len(nearby_stations), 1):
                    stations_fl.append(singleStation(nearby_stations[this_station]))
                min_dist = min(stations_fl)
            except:
                min_dist = -1
            
            print('MIN. DIST. {}.'. format(min_dist))
            
            """ ::::::::::::::::::::::::::::::::::::::::::::::::::::::::::: """            
            
            """ Trying to get Lot Lan"""
            try:
                geo_lat = blockProperty.find(itemprop = "latitude").get("content")
            except:
                geo_lat = 0
                
            try:
                geo_lon = blockProperty.find(itemprop = "longitude").get("content")
            except:
                geo_lon = 0
            print("----------------")
                
            """ get Type : working """
            finalType = ""
            typeProp = blockProperty.find("h2").find_all("a")[0].text
            if "flat".upper() in typeProp.upper():
                finalType = "flat"
            elif "semi-deteached".upper() in typeProp.upper():
                    finalType = "semi-deteached house"
            elif "property".upper() in typeProp.upper():
                finalType = "apartment"
            elif "studio".upper() in typeProp.upper():
                finalType = "studio"
            elif "maisonette".upper() in typeProp.upper():
                finalType = "maisonette"
            elif "room".upper() in typeProp.upper():
                finalType = "room"
            elif "house".upper() in typeProp.upper():
                finalType = "house"
            elif "shared accommodation".upper() in typeProp.upper():
                finalType = "shared accomodation"
                    
            """ get link """
            linkTo = str(blockProperty.find("a")).split('"/')[1].split('"')[0]
            linkTo = "https://www.zoopla.co.uk/" + linkTo
                    
            """ get full descriptiom """
            descriptionFull = getFullDescription(linkTo)
            
            """ get Floor plan """
                    
            """ Save features to Database """
            featuresOfProperties["ID"] = idPro
            featuresOfProperties["DESCRIPTION"] = descriptionText
            featuresOfProperties["STATION"] = nearbyStation
            featuresOfProperties["ADDED"] = listedOn
            featuresOfProperties["LOCATION"] = address
            featuresOfProperties['NUMBER OF CLOSEST STATIONS'] = number_of_stations
            featuresOfProperties['MIN. DISTANCE TO THE STATION'] = min_dist
            featuresOfProperties["PRICE PCM (£)"] = ppM
            featuresOfProperties["TYPE"] = finalType
            featuresOfProperties["BATHROOMS"] = bathrooms
            featuresOfProperties["BEDROOMS"] = bedrooms
            featuresOfProperties["RECEPTIONS"] = receptions
            featuresOfProperties["LINK"] = linkTo
            featuresOfProperties["FULL DESCRIPTION"] = descriptionFull
            featuresOfProperties["LON"] = geo_lon
            featuresOfProperties["LAT"] = geo_lat
                
            properties.append(dict(featuresOfProperties))
            #print("---------------------")
    return properties
""" Get Data : End"""

""" Update EXCEL file : start """
def updateExcel(dataset):
    import openpyxl, os, datetime
    from datetime import datetime
    from pandas import ExcelWriter as ewriter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill
    
    filename = "WestDrayton_Rent_20180817.xlsx" #<-- DO NOT CHANGE!!!
    wb = openpyxl.load_workbook(filename, data_only = False)          
    
    sheet = wb["Data Rents"]
    
    rowForSearching = 10
    while len(str(sheet.cell(row = rowForSearching, column = 2).value)) > 4:
        rowForSearching = rowForSearching + 1
    
    """ Define some Excel cell styles and formatting """
    fillDefault = PatternFill(fill_type = None,
                              start_color = "FFFFFFFF",
                              end_color = "FF000000")
    
    HyperlinkBlue = Font(color = "0563c1",
                         underline = "single")
    
    rightAligment = Alignment(horizontal = "right")
    centerAligment = Alignment(horizontal = "center")
    
    """ Starting writing data to Excel file """
    id_column = dataset.columns.get_loc("ID")                        # get index of ID column
    id_station = dataset.columns.get_loc("STATION")                  # get index of STATION column
    id_location = dataset.columns.get_loc("LOCATION")                # get index of LOCATION column
    id_bedrooms = dataset.columns.get_loc("BEDROOMS")                # get index of BEDROOMS column
    id_type = dataset.columns.get_loc("TYPE")                        # get index of TYPE column
    id_pricePCM = dataset.columns.get_loc("PRICE PCM (£)")           # get index of PCM column
    id_description = dataset.columns.get_loc("DESCRIPTION")          # get index of DESCRIPTION column
    id_link = dataset.columns.get_loc("LINK")                        # get index of LINK column
    id_added = dataset.columns.get_loc("ADDED")                      # get index of ADDED column
    
    dataset_index = 0
    row_index = 0
    for i in range(rowForSearching, rowForSearching + len(dataset), 1):
        """ writing MONTH column """
        today = datetime.now().strftime('%d')
        month = datetime.now().strftime('%m')
        year = "19"
        if month == "11":
            monthW = "Nov"
        elif month == "01":
            monthW = "Jan"
        elif month == "02":
            monthW = "Feb"
        elif month == "03":
            monthW = "Mar"
        elif month == "04":
            monthW = "Apr"
        elif month == "05":
            monthW = "May"
        elif month == "06":
            monthW = "Feb"
        elif month == "07":
            monthW = "Jul"
        elif month == "08":
            monthW = "Aug"   
        elif month == "09":
            monthW = "Sep"
        elif month == "10":
            monthW = "Oct"
        elif month == '12':
            monthW = 'Dec'
            
        dateNow = today + "-" + monthW + "-" + year
        sheet.cell(row = i, column = 2).value = dateNow
        sheet.cell(row = i, column = 2).alignment = rightAligment
        
        """ writing ID column """
        sheet.cell(row = i, column = 3).value = dataset.iat[dataset_index, id_column]
        print('>>> {}:{}'.format(i,
                                 dataset.iat[dataset_index, id_column]))
        
        """ writing STATION column """
        sheet.cell(row = i, column = 5).value = dataset.iat[dataset_index, id_station].upper()
        
        """ writing LOCATION column """
        sheet.cell(row = i, column = 6).value = dataset.iat[dataset_index, id_location]
        
        """ writing BEDROOMS column """
        sheet.cell(row = i, column = 7).value = dataset.iat[dataset_index, id_bedrooms]
        sheet.cell(row = i, column = 7).alignment = centerAligment
        
        """ writing TYPE column """
        sheet.cell(row = i, column = 8).value = dataset.iat[dataset_index, id_type]
        sheet.cell(row = i, column = 8).alignment = centerAligment
        
        """ writing PRICE PCM column """
        sheet.cell(row = i, column = 9).value = dataset.iat[dataset_index, id_pricePCM]
        sheet.cell(row = i, column = 9).alignment = centerAligment
        
        """ writing DESCRIPTION column """
        sheet.cell(row = i, column = 11).value = dataset.iat[dataset_index, id_description]
        
        """ writing LINK column """
        FullLink = '=HYPERLINK("' + dataset.iat[dataset_index, id_link] + '","' + "WEBPAGE" + '")'
        sheet.cell(row = i, column = 12).value = FullLink
        sheet.cell(row = i, column = 12).font = HyperlinkBlue
        sheet.cell(row = i, column = 12).alignment = centerAligment
        
        """ writing ADDED column """
        sheet.cell(row = i, column = 15).value = dataset.iat[dataset_index, id_added]
        sheet.cell(row = i, column = 15).fill = fillDefault
        
        dataset_index = dataset_index + 1
        row_index = i
        
    print('\nStarting formating.')
    print('>>> Max. row is: {}.'.format(row_index))
    for i in range(rowForSearching, row_index, 1):
        print('>>> Row: {}'.format(str(i)))
        for j in range(1, sheet.max_column, 1):
            sheet.cell(row = i, column = j).fill = fillDefault
    
    print('\nSaving Excel file.')
    """ SAVE DATA TO EXCEL """
    import datetime
    from datetime import date 
    timeNow = str(datetime.datetime.now()).replace(":","-")[:-10]
    wb.save("DataSaved-FINAL - RENT - WEST DRAYTON - " + timeNow + ".xlsx")
    
    print('Finished.')
        
    return None
""" Update EXCEL file : end """

""" Save data to file : start """
def writeToFile(DF):
    import time, datetime, os
    from datetime import date 
    
    timeNow = str(datetime.datetime.now()).replace(":","-")[:-10]
    DF.to_csv("{}_Zoopla_WEST DRAYTON.csv".format(timeNow))
    return None
""" Save data to file : end """
        
numberOfPages = 6
URLs = generateURLs(numberOfPages)
FullData = getData(URLs)
DF = pd.DataFrame(FullData)
DF = DF[["ID", "STATION", "LOCATION", "BEDROOMS", "BATHROOMS", "RECEPTIONS", "TYPE", 
         "PRICE PCM (£)", "DESCRIPTION", "LINK", "ADDED", "FULL DESCRIPTION", 
         "NUMBER OF CLOSEST STATIONS", 'MIN. DISTANCE TO THE STATION']]
DataBase = DF

# ::: Convert negatives values to the representative ones :::
max_distance = max(DataBase['MIN. DISTANCE TO THE STATION'])
theor_max_d  = max_distance * 1.5

index_of_st_dist = DataBase.columns.get_loc('MIN. DISTANCE TO THE STATION')
for this_row in range(0, len(DataBase), 1):
    if DataBase.iat[this_row, index_of_st_dist] == float(-1):
        DataBase.iat[this_row, index_of_st_dist] = theor_max_d

writeToFile(DataBase)
updateExcel(DataBase)

s = 'copy_URL_here'.encode('utf-8')